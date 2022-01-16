
import openpyxl

filename = r'cyg_deal.xlsx'
inwb = openpyxl.load_workbook(filename)


sheetname = inwb.get_sheet_names()
print(sheetname)
ws = inwb.get_sheet_by_name(sheetname[0])

rows = ws.max_row
cols = ws.max_column
#高值区域的系数
#
#中值区域的系数
# ymax = 51325.61999
# k = 233.22883
# n = 6.77031
#低值区域的系数
for i in range(2,rows+1):
    bianhao = ws.cell(i,2).value
    gray_avg = float(ws.cell(i,3).value)
    biaozhun = float(ws.cell(i,4).value)
    if gray_avg<120 :
        ymax = 1138.30363
        k = 94.63473
        n =  7.33951
    else:
        ymax = 51325.61999
        k = 233.22883
        n = 6.77031
    area = ymax * pow(gray_avg,n)/(pow(k,n)+pow(gray_avg,n))
    # print("编号：%s | 平均灰度 %f | 标准面积：%f | 公式代入得到的实际面积：%f"%(bianhao,gray_avg,biaozhun,area))
    print(area)
    

# a= 575.94708

# b = -24.33358
# c = 0.25605
# x = 130
# # result = a+b*x+c*x**2
# result = ymax * pow(x,n)/(pow(k,n)+pow(x,n))
# print(result)

# all_const = 104126.461
# v_const = 1.473
# s_const = 2.988
# h_const = -862.957


# for i in range(2,rows):
#     result = all_const+v_const*int(ws.cell(i,3).value)+s_const*int(ws.cell(i,4).value)+h_const*int(ws.cell(i,5).value)
#     print("面积%s 亮度：%s 饱和度: %s 色调：%s  公式计算的结果%f" %(ws.cell(i,2).value,ws.cell(i,3).value,ws.cell(i,4).value,ws.cell(i,5).value,result))


# ws.close()



